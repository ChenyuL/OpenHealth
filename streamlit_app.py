import streamlit as st
import requests
import json
import time
import os
from datetime import datetime
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from typing import Dict, List, Optional
import anthropic
import PyPDF2
from docx import Document
import openpyxl
import io

# Configuration
API_BASE_URL = "http://localhost:8000"
ANTHROPIC_API_KEY = "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"

# Page config
st.set_page_config(
    page_title="🏥 OpenHealth",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: linear-gradient(90deg, #f0f8ff 0%, #e6f3ff 100%);
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #1f77b4;
        margin: 0.5rem 0;
    }
    .chat-message {
        padding: 1rem;
        margin: 0.5rem 0;
        border-radius: 10px;
        border-left: 4px solid #1f77b4;
    }
    .user-message {
        background-color: #f0f8ff;
        border-left-color: #1f77b4;
    }
    .assistant-message {
        background-color: #f5f5f5;
        border-left-color: #28a745;
    }
    .thinking-message {
        background-color: #fff3cd;
        border-left-color: #ffc107;
        font-style: italic;
    }
    .sidebar-section {
        padding: 1rem 0;
        border-bottom: 1px solid #e0e0e0;
    }
    .document-card {
        background: #f8f9fa;
        padding: 0.5rem;
        margin: 0.5rem 0;
        border-radius: 5px;
        border-left: 3px solid #6c757d;
    }
</style>
""", unsafe_allow_html=True)

class AnthropicClient:
    """Anthropic API client with thinking capability"""
    
    def __init__(self, api_key: str):
        self.client = anthropic.Anthropic(api_key=api_key)
    
    def chat_with_thinking(self, message: str, conversation_history: List[Dict] = None) -> Dict:
        """Chat with Claude using thinking capability for deep research"""
        try:
            # Prepare messages
            messages = []
            if conversation_history:
                for msg in conversation_history[-10:]:  # Last 10 messages for context
                    if msg["role"] in ["user", "assistant"]:
                        messages.append({
                            "role": msg["role"],
                            "content": msg["content"]
                        })
            
            # Add current message
            messages.append({
                "role": "user",
                "content": message
            })
            
            # Call Claude with thinking enabled
            response = self.client.messages.create(
                model="claude-3-5-sonnet-20241022",  # Using latest available model
                max_tokens=16000,
                thinking={
                    "type": "enabled",
                    "budget_tokens": 10000
                },
                messages=messages
            )
            
            # Extract thinking and response
            thinking_content = ""
            response_content = ""
            
            for block in response.content:
                if block.type == "thinking":
                    thinking_content = block.thinking
                elif block.type == "text":
                    response_content = block.text
            
            return {
                "thinking": thinking_content,
                "response": response_content,
                "full_response": response
            }
            
        except Exception as e:
            return {
                "thinking": "",
                "response": f"Error calling Anthropic API: {str(e)}",
                "full_response": None
            }

class FileProcessor:
    """Process uploaded files and extract content"""
    
    @staticmethod
    def extract_pdf_content(file_bytes: bytes) -> str:
        """Extract text from PDF file"""
        try:
            pdf_file = io.BytesIO(file_bytes)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
        except Exception as e:
            return f"Error extracting PDF: {str(e)}"
    
    @staticmethod
    def extract_docx_content(file_bytes: bytes) -> str:
        """Extract text from DOCX file"""
        try:
            doc_file = io.BytesIO(file_bytes)
            doc = Document(doc_file)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            return f"Error extracting DOCX: {str(e)}"
    
    @staticmethod
    def extract_excel_content(file_bytes: bytes) -> str:
        """Extract content from Excel file"""
        try:
            excel_file = io.BytesIO(file_bytes)
            workbook = openpyxl.load_workbook(excel_file)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return text
        except Exception as e:
            return f"Error extracting Excel: {str(e)}"

class OpenHealthAPI:
    """API client for OpenHealth Agent backend"""
    
    def __init__(self, base_url: str = API_BASE_URL):
        self.base_url = base_url
    
    def get_health(self):
        """Check API health"""
        try:
            response = requests.get(f"{self.base_url}/health")
            return response.json() if response.status_code == 200 else None
        except:
            return None

# Initialize clients
if ANTHROPIC_API_KEY:
    claude_client = AnthropicClient(ANTHROPIC_API_KEY)
else:
    claude_client = None

api = OpenHealthAPI()
file_processor = FileProcessor()

# Initialize session state
if 'current_tenant_id' not in st.session_state:
    st.session_state.current_tenant_id = "demo-tenant-123"
if 'current_conversation_id' not in st.session_state:
    st.session_state.current_conversation_id = None
if 'messages' not in st.session_state:
    st.session_state.messages = []
if 'schemas' not in st.session_state:
    st.session_state.schemas = []
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []
if 'extracted_data' not in st.session_state:
    st.session_state.extracted_data = {}

def main():
    """Main application"""
    
    # Header
    st.markdown('<h1 class="main-header">🏥 OpenHealth</h1>', unsafe_allow_html=True)
    st.markdown('<p style="text-align: center; font-size: 1.2rem; color: #666;">Healthcare Venture Screening Platform with AI Intelligence</p>', unsafe_allow_html=True)
    
    # Check API health
    health = api.get_health()
    if health:
        st.success(f"✅ Backend API Connected - {health.get('status', 'healthy').title()}")
    else:
        st.warning("⚠️ Backend API not available (running in offline mode)")
    
    # Check Anthropic API
    if claude_client:
        st.success("✅ Anthropic Claude API Ready")
    else:
        st.error("❌ Anthropic API Key not configured")
    
    # Sidebar
    with st.sidebar:
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown("### 🏢 Workspace")
        
        if st.session_state.current_tenant_id:
            st.info(f"📁 Current workspace: {st.session_state.current_tenant_id[:8]}...")
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Schema selection
        st.markdown('<div class="sidebar-section">', unsafe_allow_html=True)
        st.markdown("### 📊 Analysis Schema")
        schema_options = ["Medical Device Ventures", "Digital Health Startups", "Biotech Companies", "Healthcare AI", "Pharmaceutical"]
        selected_schema = st.selectbox("Select Schema", schema_options, key="schema_select")
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Live Data Extraction Display
        if st.session_state.current_conversation_id and st.session_state.extracted_data:
            st.markdown("### 📊 Live Data Extraction")
            st.json(st.session_state.extracted_data)
    
    # Main content layout
    col1, col2 = st.columns([2, 1])
    
    with col1:
        show_ai_conversation()
    
    with col2:
        show_dashboard_metrics()
        st.markdown("---")
        show_file_upload()
        st.markdown("---")
        show_analytics_summary()

def show_ai_conversation():
    """Enhanced AI conversation interface"""
    st.header("💬 AI-Powered Healthcare Venture Analysis")
    
    # Conversation setup
    if not st.session_state.current_conversation_id:
        st.subheader("🚀 Start New Analysis Session")
        
        col1, col2 = st.columns(2)
        with col1:
            venture_name = st.text_input("Venture/CEO Name", placeholder="e.g., CardioTech - John Smith")
        with col2:
            session_title = st.text_input("Session Title", placeholder="e.g., Series A Due Diligence")
        
        if st.button("🎯 Begin Analysis", type="primary", use_container_width=True):
            if venture_name:
                st.session_state.current_conversation_id = f"session-{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                st.session_state.messages = [
                    {
                        "role": "assistant", 
                        "content": f"🔬 **Analysis Session Started for {venture_name}**\n\nI'm Claude, your AI healthcare investment analyst. I'll use deep reasoning to help you evaluate this venture comprehensively.\n\n**What I can analyze:**\n• Technology & Innovation\n• Market Opportunity\n• Regulatory Pathway\n• Team & Leadership\n• Financial Projections\n• Risk Assessment\n\nWhat would you like to explore first?", 
                        "timestamp": datetime.now(),
                        "thinking": ""
                    }
                ]
                st.session_state.extracted_data = {
                    "venture_name": venture_name,
                    "session_title": session_title or "Analysis Session",
                    "started": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "schema": st.session_state.get("schema_select", "Medical Device Ventures"),
                    "completion": "5%"
                }
                st.rerun()
    
    # Chat interface
    if st.session_state.current_conversation_id:
        st.subheader(f"💬 Active Session: {st.session_state.extracted_data.get('venture_name', 'Analysis')}")
        
        # Messages container
        messages_container = st.container(height=500)
        
        with messages_container:
            for message in st.session_state.messages:
                role = message["role"]
                content = message["content"]
                timestamp = message.get("timestamp", datetime.now())
                thinking = message.get("thinking", "")
                
                if role == "user":
                    st.markdown(f'''
                    <div class="chat-message user-message">
                        <strong>👤 You</strong> <small>({timestamp.strftime("%H:%M")})</small><br>
                        {content}
                    </div>
                    ''', unsafe_allow_html=True)
                else:
                    # Show thinking process if available
                    if thinking:
                        with st.expander("🧠 AI Thinking Process", expanded=False):
                            st.markdown(f'''
                            <div class="thinking-message">
                                <strong>💭 Claude's Analysis Process:</strong><br>
                                {thinking[:500]}{'...' if len(thinking) > 500 else ''}
                            </div>
                            ''', unsafe_allow_html=True)
                    
                    st.markdown(f'''
                    <div class="chat-message assistant-message">
                        <strong>🤖 Claude AI</strong> <small>({timestamp.strftime("%H:%M")})</small><br>
                        {content}
                    </div>
                    ''', unsafe_allow_html=True)
        
        # Message input
        st.markdown("### 💭 Your Question")
        user_input = st.text_area(
            "Ask about the venture...", 
            placeholder="Examples:\n• What are the key regulatory risks for this medical device?\n• How does their technology compare to competitors?\n• What's your assessment of the market opportunity?",
            height=100,
            key="user_input"
        )
        
        col1, col2, col3 = st.columns([2, 1, 1])
        
        with col1:
            if st.button("📤 Send Message", type="primary", disabled=not claude_client):
                if user_input and claude_client:
                    # Add user message
                    st.session_state.messages.append({
                        "role": "user", 
                        "content": user_input, 
                        "timestamp": datetime.now(),
                        "thinking": ""
                    })
                    
                    # Show thinking indicator
                    with st.spinner("🧠 Claude is thinking deeply about your question..."):
                        # Get AI response with thinking
                        ai_result = claude_client.chat_with_thinking(
                            user_input, 
                            st.session_state.messages
                        )
                        
                        # Add AI response
                        st.session_state.messages.append({
                            "role": "assistant", 
                            "content": ai_result["response"], 
                            "timestamp": datetime.now(),
                            "thinking": ai_result["thinking"]
                        })
                        
                        # Update extracted data based on conversation
                        update_extracted_data(user_input, ai_result["response"])
                    
                    # Clear input and rerun
                    st.session_state.user_input = ""
                    st.rerun()
        
        with col2:
            if st.button("🎯 Suggest Questions"):
                questions = get_suggested_questions(st.session_state.get("schema_select", "Medical Device Ventures"))
                suggestion_text = "**Suggested questions for " + st.session_state.get("schema_select", "Medical Device Ventures") + ":**\n\n"
                suggestion_text += "\n".join([f"• {q}" for q in questions[:5]])
                
                st.session_state.messages.append({
                    "role": "assistant", 
                    "content": suggestion_text, 
                    "timestamp": datetime.now(),
                    "thinking": ""
                })
                st.rerun()
        
        with col3:
            if st.button("📋 Generate Summary"):
                if claude_client:
                    with st.spinner("📝 Generating comprehensive analysis summary..."):
                        summary_prompt = f"Based on our conversation about {st.session_state.extracted_data.get('venture_name', 'this venture')}, please provide a comprehensive investment analysis summary covering: 1) Technology Assessment, 2) Market Opportunity, 3) Regulatory Pathway, 4) Team Evaluation, 5) Key Risks, 6) Investment Recommendation with reasoning."
                        
                        summary_result = claude_client.chat_with_thinking(
                            summary_prompt,
                            st.session_state.messages
                        )
                        
                        st.session_state.messages.append({
                            "role": "assistant", 
                            "content": f"📊 **Comprehensive Investment Analysis Summary**\n\n{summary_result['response']}", 
                            "timestamp": datetime.now(),
                            "thinking": summary_result["thinking"]
                        })
                        st.rerun()
        
        # Reset conversation
        if st.button("🔄 Start New Session", type="secondary"):
            st.session_state.current_conversation_id = None
            st.session_state.messages = []
            st.session_state.extracted_data = {}
            st.rerun()

def show_file_upload():
    """File upload and processing"""
    st.subheader("📁 Document Analysis")
    
    # File uploader
    uploaded_files = st.file_uploader(
        "Upload venture documents",
        accept_multiple_files=True,
        type=['pdf', 'docx', 'xlsx'],
        help="Upload pitch decks, financial models, technical documents, etc."
    )
    
    if uploaded_files:
        st.markdown("**📄 Uploaded Files:**")
        
        for i, file in enumerate(uploaded_files):
            col1, col2 = st.columns([3, 1])
            
            with col1:
                st.markdown(f'''
                <div class="document-card">
                    <strong>📄 {file.name}</strong><br>
                    <small>Type: {file.type} | Size: {file.size:,} bytes</small>
                </div>
                ''', unsafe_allow_html=True)
            
            with col2:
                if st.button("🔍 Analyze", key=f"analyze_{i}"):
                    with st.spinner(f"Processing {file.name}..."):
                        # Read file content
                        file_bytes = file.read()
                        
                        # Extract content based on file type
                        if file.type == "application/pdf":
                            content = file_processor.extract_pdf_content(file_bytes)
                        elif file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                            content = file_processor.extract_docx_content(file_bytes)
                        elif file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                            content = file_processor.extract_excel_content(file_bytes)
                        else:
                            content = "Unsupported file type"
                        
                        # Add file content to conversation for analysis
                        if claude_client and content and not content.startswith("Error"):
                            file_analysis_prompt = f"Please analyze this document content from {file.name} and extract key healthcare venture information relevant to investment analysis:\n\n{content[:5000]}..."
                            
                            analysis_result = claude_client.chat_with_thinking(
                                file_analysis_prompt,
                                st.session_state.messages
                            )
                            
                            # Add to conversation
                            st.session_state.messages.append({
                                "role": "assistant",
                                "content": f"📄 **Document Analysis: {file.name}**\n\n{analysis_result['response']}",
                                "timestamp": datetime.now(),
                                "thinking": analysis_result["thinking"]
                            })
                            
                            # Store file in session
                            st.session_state.uploaded_files.append({
                                "name": file.name,
                                "content": content,
                                "analysis": analysis_result['response'],
                                "uploaded": datetime.now()
                            })
                            
                            st.success(f"✅ {file.name} analyzed successfully!")
                            st.rerun()
                        else:
                            st.error(f"❌ Could not process {file.name}")

def show_dashboard_metrics():
    """Dashboard metrics summary"""
    st.subheader("📊 Analysis Overview")
    
    # Metrics
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric("Active Sessions", "1" if st.session_state.current_conversation_id else "0")
        st.metric("Documents", len(st.session_state.uploaded_files))
    
    with col2:
        completion = st.session_state.extracted_data.get("completion", "0%")
        st.metric("Analysis Progress", completion)
        st.metric("AI Interactions", len([m for m in st.session_state.messages if m["role"] == "assistant"]))

def show_analytics_summary():
    """Analytics and insights summary"""
    st.subheader("📈 Quick Insights")
    
    if st.session_state.current_conversation_id:
        # Mock analytics based on conversation
        venture_name = st.session_state.extracted_data.get("venture_name", "Current Venture")
        
        # Risk assessment gauge
        risk_score = 65  # Mock score
        fig = go.Figure(go.Indicator(
            mode = "gauge+number",
            value = risk_score,
            domain = {'x': [0, 1], 'y': [0, 1]},
            title = {'text': "Risk Score"},
            gauge = {
                'axis': {'range': [None, 100]},
                'bar': {'color': "darkblue"},
                'steps': [
                    {'range': [0, 50], 'color': "lightgray"},
                    {'range': [50, 80], 'color': "gray"}],
                'threshold': {
                    'line': {'color': "red", 'width': 4},
                    'thickness': 0.75,
                    'value': 90}}))
        
        fig.update_layout(height=250)
        st.plotly_chart(fig, use_container_width=True)
        
        # Key metrics
        if st.session_state.extracted_data:
            st.markdown("**Key Findings:**")
            for key, value in st.session_state.extracted_data.items():
                if key not in ["venture_name", "session_title", "started"]:
                    st.write(f"• **{key.replace('_', ' ').title()}**: {value}")
    else:
        st.info("Start an analysis session to see insights")

def update_extracted_data(user_question: str, ai_response: str):
    """Update extracted data based on conversation"""
    # Simple keyword-based extraction (in real app, this would be more sophisticated)
    
    # Update completion percentage
    current_completion = int(st.session_state.extracted_data.get("completion", "0%").replace("%", ""))
    new_completion = min(current_completion + 15, 95)
    st.session_state.extracted_data["completion"] = f"{new_completion}%"
    
    # Extract key information based on keywords
    if any(word in user_question.lower() for word in ["market", "size", "opportunity"]):
        if "market" not in st.session_state.extracted_data:
            st.session_state.extracted_data["market_size"] = "Under Analysis"
    
    if any(word in user_question.lower() for word in ["regulatory", "fda", "approval"]):
        if "regulatory" not in st.session_state.extracted_data:
            st.session_state.extracted_data["regulatory_pathway"] = "Under Review"
    
    if any(word in user_question.lower() for word in ["technology", "innovation", "tech"]):
        if "technology" not in st.session_state.extracted_data:
            st.session_state.extracted_data["technology_assessment"] = "In Progress"
    
    if any(word in user_question.lower() for word in ["team", "founder", "leadership"]):
        if "team" not in st.session_state.extracted_data:
            st.session_state.extracted_data["team_evaluation"] = "Analyzed"

def get_suggested_questions(schema: str) -> List[str]:
    """Get suggested questions based on schema type"""
    questions_map = {
        "Medical Device Ventures": [
            "What is the device classification and regulatory pathway?",
            "What clinical data supports the efficacy claims?",
            "Who are the key competitors and how does this differentiate?",
            "What is the reimbursement strategy and payer coverage?",
            "What intellectual property protection exists?",
            "What are the manufacturing and quality control processes?",
            "How scalable is the technology platform?"
        ],
        "Digital Health Startups": [
            "How does the platform ensure data privacy and security?",
            "What is the user engagement and retention data?",
            "How will you achieve clinical validation?",
            "What is the business model and revenue streams?",
            "How do you plan to scale user acquisition?",
            "What partnerships exist with healthcare providers?",
            "How does the solution integrate with existing workflows?"
        ],
        "Biotech Companies": [
            "What stage are the lead compounds in development?",
            "What is the mechanism of action and competitive advantage?",
            "What clinical trial data is available?",
            "What is the intellectual property landscape?",
            "What partnerships exist with pharma companies?",
            "What are the manufacturing and scale-up plans?",
            "How does this address unmet medical needs?"
        ],
        "Healthcare AI": [
            "What training data was used for the AI models?",
            "How do you ensure algorithmic fairness and bias mitigation?",
            "What validation studies demonstrate clinical utility?",
            "How will the AI integrate with existing clinical workflows?",
            "What is the accuracy compared to current standard of care?",
            "How do you handle explainability and transparency?",
            "What are the data governance and privacy measures?"
        ],
        "Pharmaceutical": [
            "What is the drug development pipeline status?",
            "What are the target indications and patient populations?",
            "What clinical trial results support efficacy?",
            "What is the competitive landscape analysis?",
            "What regulatory approval strategy is planned?",
            "What are the manufacturing and supply chain plans?",
            "How does this address current treatment gaps?"
        ]
    }
    
    return questions_map.get(schema, questions_map["Medical Device Ventures"])

if __name__ == "__main__":
    main()
